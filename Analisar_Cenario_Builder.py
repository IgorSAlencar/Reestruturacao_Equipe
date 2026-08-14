#!/usr/bin/env python3
"""Analisa e compara um cenário exportado pelo Builder com o cenário atual.

O programa aceita tanto o JSON bruto de ``ScenarioData`` quanto o envelope de
rascunho exportado pela aplicação (com o cenário dentro da chave ``data``).
As distâncias são sempre recalculadas por Haversine a partir das coordenadas.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd

EARTH_RADIUS_KM = 6_371.0088
METRIC_FIELDS = (
    "municipios",
    "distritos",
    "unidades_territoriais",
    "correspondentes",
    "populacao",
    "km_min",
    "km_medio",
    "km_mediano",
    "km_p90",
    "km_max",
    "km_medio_ponderado_correspondentes",
    "km_medio_ponderado_populacao",
)


@dataclass
class PreparedScenario:
    name: str
    source: str
    raw: dict[str, Any]
    poles: dict[str, dict[str, Any]]
    rows: list[dict[str, Any]]
    issues: list[dict[str, Any]]


@dataclass
class AnalysisReport:
    current: PreparedScenario
    proposed: PreparedScenario
    summary: pd.DataFrame
    current_portfolios: pd.DataFrame
    proposed_portfolios: pd.DataFrame
    portfolio_comparison: pd.DataFrame
    area_comparison: pd.DataFrame
    movements: pd.DataFrame
    unmatched: pd.DataFrame
    quality: pd.DataFrame
    insights: pd.DataFrame


def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calcula a distância geodésica em quilômetros entre dois pontos."""
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    return EARTH_RADIUS_KM * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def _normalise_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).lower()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return re.sub(r"\.0$", "", text)


def _safe_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def _valid_lat_lon(lat: float | None, lon: float | None) -> bool:
    return lat is not None and lon is not None and -90 <= lat <= 90 and -180 <= lon <= 180


def _coordinate(item: Mapping[str, Any], short_name: str, long_name: str) -> float | None:
    value = item.get(short_name)
    if value is None or value == "":
        value = item.get(long_name)
    return _safe_float(value)


def _stable_id_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value.casefold())


def _issue(
    issues: list[dict[str, Any]],
    scenario: str,
    severity: str,
    code: str,
    entity_type: str,
    entity_id: str,
    message: str,
) -> None:
    issues.append(
        {
            "cenario": scenario,
            "severidade": severity,
            "codigo": code,
            "tipo_entidade": entity_type,
            "entidade_id": entity_id,
            "mensagem": message,
        }
    )


def unwrap_scenario(payload: Mapping[str, Any]) -> dict[str, Any]:
    """Extrai ``ScenarioData`` de um envelope do Builder ou retorna o objeto bruto."""
    if not isinstance(payload, Mapping):
        raise TypeError("O JSON precisa conter um objeto na raiz.")
    nested = payload.get("data")
    if isinstance(nested, Mapping) and ("poles" in nested or "units" in nested):
        return dict(nested)
    return dict(payload)


def load_scenario(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    try:
        with source.open("r", encoding="utf-8-sig") as handle:
            payload = json.load(handle)
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON inválido em {source}: {exc}") from exc
    return unwrap_scenario(payload)


def _coverage_keys(unit: Mapping[str, Any]) -> tuple[str, str, str]:
    unit_type = str(unit.get("type") or "MUNICIPALITY").strip().upper()
    municipality_code = _normalise_id(unit.get("municipalityCode"))
    if not municipality_code:
        fallback = f"{unit.get('uf', '')}:{unit.get('name', '')}".strip(":")
        municipality_code = re.sub(r"\s+", "_", fallback.casefold()) or _normalise_id(unit.get("id"))
    municipal_key = f"MUN:{municipality_code}"
    if unit_type in {"DISTRICT", "DISTRITO"}:
        district_code = _normalise_id(unit.get("districtCode")) or _normalise_id(unit.get("id"))
        exact_key = f"DIST:{district_code}"
        canonical_type = "DISTRITO"
    else:
        exact_key = municipal_key
        canonical_type = "MUNICIPIO"
    return exact_key, municipal_key, canonical_type


def prepare_scenario(
    payload: Mapping[str, Any],
    label: str,
    source: str = "",
) -> PreparedScenario:
    """Valida e transforma um cenário em unidades territoriais canônicas."""
    data = unwrap_scenario(payload)
    issues: list[dict[str, Any]] = []
    raw_poles = data.get("poles")
    raw_units = data.get("units")
    if not isinstance(raw_poles, list):
        raise TypeError(f"{label}: a chave 'poles' precisa ser uma lista.")
    if not isinstance(raw_units, list):
        raise TypeError(f"{label}: a chave 'units' precisa ser uma lista.")
    if not isinstance(data.get("summary"), Mapping):
        _issue(issues, label, "AVISO", "SUMMARY_AUSENTE", "CENARIO", "", "Resumo ausente ou inválido.")

    poles: dict[str, dict[str, Any]] = {}
    for index, item in enumerate(raw_poles):
        if not isinstance(item, Mapping):
            _issue(issues, label, "ERRO", "POLO_INVALIDO", "POLO", str(index), "Registro de polo não é um objeto.")
            continue
        pole = dict(item)
        pole_id = _normalise_id(pole.get("id"))
        if not pole_id:
            _issue(issues, label, "ERRO", "POLO_SEM_ID", "POLO", str(index), "Polo sem identificador; registro ignorado.")
            continue
        if pole_id in poles:
            _issue(issues, label, "ERRO", "POLO_DUPLICADO", "POLO", pole_id, "ID de polo duplicado; mantido o primeiro registro.")
            continue
        pole["id"] = pole_id
        pole["lat"] = _coordinate(pole, "lat", "latitude")
        pole["lon"] = _coordinate(pole, "lon", "longitude")
        pole["area"] = str(pole.get("area") or "SEM GERENCIA DE AREA").strip()
        pole["name"] = str(pole.get("name") or f"Polo {pole_id}").strip()
        if not _valid_lat_lon(pole["lat"], pole["lon"]):
            _issue(issues, label, "ERRO", "COORDENADA_POLO_INVALIDA", "POLO", pole_id, "Latitude/longitude do polo ausente ou fora do intervalo válido.")
        poles[pole_id] = pole

    seen_unit_ids: set[str] = set()
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    for index, item in enumerate(raw_units):
        if not isinstance(item, Mapping):
            _issue(issues, label, "ERRO", "UNIDADE_INVALIDA", "UNIDADE", str(index), "Registro de unidade não é um objeto.")
            continue
        unit = dict(item)
        unit_id = _normalise_id(unit.get("id")) or f"linha-{index + 1}"
        if unit_id in seen_unit_ids:
            _issue(issues, label, "AVISO", "UNIDADE_ID_DUPLICADO", "UNIDADE", unit_id, "ID de unidade repetido; o registro ainda participa da consolidação.")
        seen_unit_ids.add(unit_id)
        pole_id = _normalise_id(unit.get("poleId"))
        if not pole_id:
            _issue(issues, label, "ERRO", "UNIDADE_SEM_POLO", "UNIDADE", unit_id, "Unidade sem poleId; não entra nas métricas de carteira.")
        elif pole_id not in poles:
            _issue(issues, label, "ERRO", "POLO_NAO_ENCONTRADO", "UNIDADE", unit_id, f"poleId {pole_id} não existe na lista de polos.")

        lat = _coordinate(unit, "lat", "latitude")
        lon = _coordinate(unit, "lon", "longitude")
        if not _valid_lat_lon(lat, lon):
            _issue(issues, label, "ERRO", "COORDENADA_UNIDADE_INVALIDA", "UNIDADE", unit_id, "Latitude/longitude da unidade ausente ou fora do intervalo válido.")
        stores = _safe_float(unit.get("stores"))
        population = _safe_float(unit.get("population"))
        if stores is None or stores < 0:
            _issue(issues, label, "AVISO", "CORRESPONDENTES_INVALIDOS", "UNIDADE", unit_id, "Quantidade de correspondentes inválida; considerado zero.")
            stores = 0.0
        if population is None or population < 0:
            _issue(issues, label, "AVISO", "POPULACAO_INVALIDA", "UNIDADE", unit_id, "População inválida; considerada zero.")
            population = 0.0

        exact_key, municipal_key, unit_type = _coverage_keys(unit)
        group_key = (pole_id, exact_key)
        current = groups.get(group_key)
        if current is None:
            groups[group_key] = {
                "pole_id": pole_id,
                "unidade_chave": exact_key,
                "cobertura_municipal_chave": municipal_key,
                "tipo": unit_type,
                "municipio_codigo": municipal_key.removeprefix("MUN:"),
                "distrito_codigo": exact_key.removeprefix("DIST:") if unit_type == "DISTRITO" else "",
                "nome": str(
                    unit.get("name")
                    or unit.get("districtName")
                    or unit.get("municipalityName")
                    or exact_key
                ),
                "uf": str(unit.get("uf") or ""),
                "latitude": lat,
                "longitude": lon,
                "correspondentes": float(stores),
                "populacao": float(population),
                "ids_origem": [unit_id],
                "registros_consolidados": 1,
            }
        else:
            current["correspondentes"] += float(stores)
            current["populacao"] = max(current["populacao"], float(population))
            current["ids_origem"].append(unit_id)
            current["registros_consolidados"] += 1
            if _valid_lat_lon(lat, lon) and _valid_lat_lon(current["latitude"], current["longitude"]):
                if abs(lat - current["latitude"]) > 1e-6 or abs(lon - current["longitude"]) > 1e-6:
                    _issue(issues, label, "AVISO", "COORDENADAS_DIVERGENTES", "UNIDADE", unit_id, "Unidades consolidadas têm coordenadas diferentes; mantida a primeira.")
            elif _valid_lat_lon(lat, lon):
                current["latitude"], current["longitude"] = lat, lon

    rows: list[dict[str, Any]] = []
    for (pole_id, _), row in sorted(groups.items(), key=lambda item: (_stable_id_key(item[0][0]), item[0][1])):
        if pole_id not in poles:
            continue
        pole = poles[pole_id]
        row["polo"] = pole["name"]
        row["gerencia_area"] = pole["area"]
        row["regional"] = str(pole.get("regional") or "")
        if _valid_lat_lon(row["latitude"], row["longitude"]) and _valid_lat_lon(pole["lat"], pole["lon"]):
            row["distancia_km"] = haversine_km(pole["lat"], pole["lon"], row["latitude"], row["longitude"])
        else:
            row["distancia_km"] = math.nan
        rows.append(row)

    summary = data.get("summary") if isinstance(data.get("summary"), Mapping) else {}
    expected_poles = _safe_float(summary.get("totalPoles"))
    expected_units = _safe_float(summary.get("totalUnits"))
    if expected_poles is not None and int(expected_poles) != len(raw_poles):
        _issue(issues, label, "AVISO", "SUMMARY_POLOS_DIVERGENTE", "CENARIO", "", f"summary.totalPoles={int(expected_poles)}, mas há {len(raw_poles)} registros.")
    if expected_units is not None and int(expected_units) != len(raw_units):
        _issue(issues, label, "AVISO", "SUMMARY_UNIDADES_DIVERGENTE", "CENARIO", "", f"summary.totalUnits={int(expected_units)}, mas há {len(raw_units)} registros.")

    name = str(data.get("title") or data.get("name") or label)
    return PreparedScenario(name=name, source=source, raw=data, poles=poles, rows=rows, issues=issues)


def _percentile(values: Sequence[float], percentile: float) -> float:
    ordered = sorted(values)
    if not ordered:
        return math.nan
    position = (len(ordered) - 1) * percentile
    lower, upper = math.floor(position), math.ceil(position)
    if lower == upper:
        return ordered[lower]
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (position - lower)


def _weighted_mean(rows: Sequence[Mapping[str, Any]], weight_field: str) -> float:
    valid = [row for row in rows if math.isfinite(float(row["distancia_km"])) and float(row[weight_field]) > 0]
    denominator = sum(float(row[weight_field]) for row in valid)
    if denominator <= 0:
        return math.nan
    return sum(float(row["distancia_km"]) * float(row[weight_field]) for row in valid) / denominator


def calculate_metrics(rows: Sequence[Mapping[str, Any]]) -> dict[str, float | int]:
    distances = [float(row["distancia_km"]) for row in rows if math.isfinite(float(row["distancia_km"]))]
    municipality_codes = {str(row["municipio_codigo"]) for row in rows if row["tipo"] == "MUNICIPIO"}
    district_codes = {str(row["distrito_codigo"]) for row in rows if row["tipo"] == "DISTRITO"}
    return {
        "municipios": len(municipality_codes),
        "distritos": len(district_codes),
        "unidades_territoriais": len(rows),
        "correspondentes": sum(float(row["correspondentes"]) for row in rows),
        "populacao": sum(float(row["populacao"]) for row in rows),
        "km_min": min(distances) if distances else math.nan,
        "km_medio": sum(distances) / len(distances) if distances else math.nan,
        "km_mediano": _percentile(distances, 0.5),
        "km_p90": _percentile(distances, 0.9),
        "km_max": max(distances) if distances else math.nan,
        "km_medio_ponderado_correspondentes": _weighted_mean(rows, "correspondentes"),
        "km_medio_ponderado_populacao": _weighted_mean(rows, "populacao"),
    }


def _portfolio_frame(scenario: PreparedScenario, label: str) -> pd.DataFrame:
    by_pole: dict[str, list[dict[str, Any]]] = {pole_id: [] for pole_id in scenario.poles}
    for row in scenario.rows:
        by_pole[row["pole_id"]].append(row)
    records: list[dict[str, Any]] = []
    for pole_id in sorted(scenario.poles, key=_stable_id_key):
        pole = scenario.poles[pole_id]
        records.append(
            {
                "cenario": label,
                "pole_id": pole_id,
                "polo": pole["name"],
                "gerencia_area": pole["area"],
                "regional": str(pole.get("regional") or ""),
                "uf": str(pole.get("uf") or ""),
                **calculate_metrics(by_pole[pole_id]),
            }
        )
    return pd.DataFrame(records)


def _delta(current: float, proposed: float) -> tuple[float, float]:
    try:
        old, new = float(current), float(proposed)
    except (TypeError, ValueError):
        return math.nan, math.nan
    if not math.isfinite(old) or not math.isfinite(new):
        return math.nan, math.nan
    absolute = new - old
    return absolute, absolute / old if old != 0 else math.nan


def _portfolio_comparison(
    current: PreparedScenario,
    proposed: PreparedScenario,
    current_frame: pd.DataFrame,
    proposed_frame: pd.DataFrame,
) -> pd.DataFrame:
    current_index = current_frame.set_index("pole_id").to_dict("index") if not current_frame.empty else {}
    proposed_index = proposed_frame.set_index("pole_id").to_dict("index") if not proposed_frame.empty else {}
    current_rows: dict[str, list[dict[str, Any]]] = {}
    proposed_rows: dict[str, list[dict[str, Any]]] = {}
    for row in current.rows:
        current_rows.setdefault(row["pole_id"], []).append(row)
    for row in proposed.rows:
        proposed_rows.setdefault(row["pole_id"], []).append(row)

    records: list[dict[str, Any]] = []
    for pole_id in sorted(set(current_index) & set(proposed_index), key=_stable_id_key):
        old, new = current_index[pole_id], proposed_index[pole_id]
        old_exact = {row["unidade_chave"] for row in current_rows.get(pole_id, [])}
        new_exact = {row["unidade_chave"] for row in proposed_rows.get(pole_id, [])}
        old_municipal = {row["cobertura_municipal_chave"] for row in current_rows.get(pole_id, [])}
        new_municipal = {row["cobertura_municipal_chave"] for row in proposed_rows.get(pole_id, [])}
        retained = old_exact & new_exact
        union = old_exact | new_exact
        municipal_retained = old_municipal & new_municipal
        record: dict[str, Any] = {
            "pole_id": pole_id,
            "polo_atual": old["polo"],
            "polo_proposto": new["polo"],
            "gerencia_area_atual": old["gerencia_area"],
            "gerencia_area_proposta": new["gerencia_area"],
        }
        for metric in METRIC_FIELDS:
            record[f"atual_{metric}"] = old[metric]
            record[f"proposto_{metric}"] = new[metric]
            record[f"delta_{metric}"], record[f"delta_pct_{metric}"] = _delta(old[metric], new[metric])
        record.update(
            {
                "unidades_retidas": len(retained),
                "unidades_removidas": len(old_exact - new_exact),
                "unidades_adicionadas": len(new_exact - old_exact),
                "sobreposicao_exata_pct": len(retained) / len(union) if union else math.nan,
                "retencao_exata_pct": len(retained) / len(old_exact) if old_exact else math.nan,
                "retencao_cobertura_municipal_pct": len(municipal_retained) / len(old_municipal) if old_municipal else math.nan,
            }
        )
        records.append(record)
    return pd.DataFrame(records)


def _area_frame(scenario: PreparedScenario) -> dict[str, dict[str, Any]]:
    by_area: dict[str, list[dict[str, Any]]] = {}
    for row in scenario.rows:
        by_area.setdefault(row["gerencia_area"], []).append(row)
    pole_counts: dict[str, int] = {}
    for pole in scenario.poles.values():
        pole_counts[pole["area"]] = pole_counts.get(pole["area"], 0) + 1
    return {
        area: {"carteiras": pole_counts.get(area, 0), **calculate_metrics(by_area.get(area, []))}
        for area in sorted(set(by_area) | set(pole_counts))
    }


def _area_comparison(current: PreparedScenario, proposed: PreparedScenario) -> pd.DataFrame:
    old_areas, new_areas = _area_frame(current), _area_frame(proposed)
    records: list[dict[str, Any]] = []
    for area in sorted(set(old_areas) | set(new_areas)):
        old = old_areas.get(area, {"carteiras": 0, **calculate_metrics([])})
        new = new_areas.get(area, {"carteiras": 0, **calculate_metrics([])})
        record: dict[str, Any] = {"gerencia_area": area}
        for metric in ("carteiras", *METRIC_FIELDS):
            record[f"atual_{metric}"] = old[metric]
            record[f"proposto_{metric}"] = new[metric]
            record[f"delta_{metric}"], record[f"delta_pct_{metric}"] = _delta(old[metric], new[metric])
        records.append(record)
    return pd.DataFrame(records)


def _dominant_current(rows: Sequence[Mapping[str, Any]]) -> tuple[str, float]:
    stores: dict[str, float] = {}
    for row in rows:
        stores[row["pole_id"]] = stores.get(row["pole_id"], 0.0) + float(row["correspondentes"])
    if not stores:
        return "", 0.0
    pole_id = min(stores, key=lambda value: (-stores[value], _stable_id_key(value)))
    return pole_id, stores[pole_id]


def _distance_for_origin(rows: Sequence[Mapping[str, Any]], pole_id: str) -> float:
    matches = [row for row in rows if row["pole_id"] == pole_id and math.isfinite(float(row["distancia_km"]))]
    if not matches:
        return math.nan
    weight = sum(float(row["correspondentes"]) for row in matches)
    if weight > 0:
        return sum(float(row["distancia_km"]) * float(row["correspondentes"]) for row in matches) / weight
    return sum(float(row["distancia_km"]) for row in matches) / len(matches)


def _movements(current: PreparedScenario, proposed: PreparedScenario) -> pd.DataFrame:
    old_by_coverage: dict[str, list[dict[str, Any]]] = {}
    new_by_coverage: dict[str, list[dict[str, Any]]] = {}
    for row in current.rows:
        old_by_coverage.setdefault(row["cobertura_municipal_chave"], []).append(row)
    for row in proposed.rows:
        new_by_coverage.setdefault(row["cobertura_municipal_chave"], []).append(row)

    records: list[dict[str, Any]] = []
    for new in proposed.rows:
        candidates = old_by_coverage.get(new["cobertura_municipal_chave"], [])
        origin_id, origin_stores = _dominant_current(candidates)
        destination_id = new["pole_id"]
        if not origin_id:
            movement = "NOVA_COBERTURA"
        elif origin_id == destination_id:
            movement = "MANTIDA"
        else:
            movement = "MOVIDA"
        origin_pole = current.poles.get(origin_id, {})
        destination_pole = proposed.poles.get(destination_id, {})
        old_distance = _distance_for_origin(candidates, origin_id) if origin_id else math.nan
        new_distance = float(new["distancia_km"])
        records.append(
            {
                "tipo_movimento": movement,
                "unidade_chave": new["unidade_chave"],
                "cobertura_municipal_chave": new["cobertura_municipal_chave"],
                "tipo_unidade": new["tipo"],
                "unidade": new["nome"],
                "uf": new["uf"],
                "polo_origem_id": origin_id,
                "polo_origem": origin_pole.get("name", ""),
                "gerencia_area_origem": origin_pole.get("area", ""),
                "polo_destino_id": destination_id,
                "polo_destino": destination_pole.get("name", ""),
                "gerencia_area_destino": destination_pole.get("area", ""),
                "correspondentes_origem_dominante": origin_stores,
                "correspondentes_propostos": new["correspondentes"],
                "populacao_proposta": new["populacao"],
                "km_atual": old_distance,
                "km_proposto": new_distance,
                "delta_km": new_distance - old_distance if math.isfinite(old_distance) and math.isfinite(new_distance) else math.nan,
            }
        )

    for coverage in sorted(set(old_by_coverage) - set(new_by_coverage)):
        old_rows = old_by_coverage[coverage]
        origin_id, origin_stores = _dominant_current(old_rows)
        origin_pole = current.poles.get(origin_id, {})
        representative = min(old_rows, key=lambda row: row["unidade_chave"])
        records.append(
            {
                "tipo_movimento": "REMOVIDA",
                "unidade_chave": representative["unidade_chave"],
                "cobertura_municipal_chave": coverage,
                "tipo_unidade": representative["tipo"],
                "unidade": representative["nome"],
                "uf": representative["uf"],
                "polo_origem_id": origin_id,
                "polo_origem": origin_pole.get("name", ""),
                "gerencia_area_origem": origin_pole.get("area", ""),
                "polo_destino_id": "",
                "polo_destino": "",
                "gerencia_area_destino": "",
                "correspondentes_origem_dominante": origin_stores,
                "correspondentes_propostos": 0.0,
                "populacao_proposta": 0.0,
                "km_atual": _distance_for_origin(old_rows, origin_id),
                "km_proposto": math.nan,
                "delta_km": math.nan,
            }
        )
    return pd.DataFrame(records)


def _unmatched(
    current: PreparedScenario,
    proposed: PreparedScenario,
    current_frame: pd.DataFrame,
    proposed_frame: pd.DataFrame,
) -> pd.DataFrame:
    old_ids, new_ids = set(current.poles), set(proposed.poles)
    records: list[dict[str, Any]] = []
    for label, ids, frame in (
        ("SOMENTE_ATUAL", old_ids - new_ids, current_frame),
        ("SOMENTE_PROPOSTO", new_ids - old_ids, proposed_frame),
    ):
        index = frame.set_index("pole_id").to_dict("index") if not frame.empty else {}
        for pole_id in sorted(ids, key=_stable_id_key):
            row = index[pole_id]
            records.append({"situacao": label, "pole_id": pole_id, **row})
    columns = ["situacao", *current_frame.columns.tolist()]
    return pd.DataFrame(records, columns=columns)


def _summary(current: PreparedScenario, proposed: PreparedScenario) -> pd.DataFrame:
    old, new = calculate_metrics(current.rows), calculate_metrics(proposed.rows)
    indicators: list[tuple[str, float | int, float | int]] = [
        ("Carteiras", len(current.poles), len(proposed.poles)),
        ("Municípios", old["municipios"], new["municipios"]),
        ("Distritos", old["distritos"], new["distritos"]),
        ("Unidades territoriais", old["unidades_territoriais"], new["unidades_territoriais"]),
        ("Correspondentes", old["correspondentes"], new["correspondentes"]),
        ("População", old["populacao"], new["populacao"]),
        ("Km médio", old["km_medio"], new["km_medio"]),
        ("Km mediano", old["km_mediano"], new["km_mediano"]),
        ("Km P90", old["km_p90"], new["km_p90"]),
        ("Km máximo", old["km_max"], new["km_max"]),
        ("Km médio ponderado por correspondentes", old["km_medio_ponderado_correspondentes"], new["km_medio_ponderado_correspondentes"]),
        ("Km médio ponderado por população", old["km_medio_ponderado_populacao"], new["km_medio_ponderado_populacao"]),
    ]
    records = []
    for indicator, old_value, new_value in indicators:
        absolute, percentage = _delta(old_value, new_value)
        records.append({"indicador": indicator, "atual": old_value, "proposto": new_value, "delta": absolute, "delta_pct": percentage})
    return pd.DataFrame(records)


def _fmt_number(value: Any, decimals: int = 1) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "n/d"
    if not math.isfinite(number):
        return "n/d"
    return f"{number:,.{decimals}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _build_insights(
    summary: pd.DataFrame,
    comparison: pd.DataFrame,
    area: pd.DataFrame,
    unmatched: pd.DataFrame,
    quality: pd.DataFrame,
) -> pd.DataFrame:
    records: list[dict[str, str]] = []

    def add(priority: str, theme: str, insight: str, evidence: str) -> None:
        records.append({"prioridade": priority, "tema": theme, "insight": insight, "evidencia": evidence})

    summary_index = summary.set_index("indicador")
    km = summary_index.loc["Km médio"]
    stores = summary_index.loc["Correspondentes"]
    km_delta = float(km["delta"])
    if km_delta > 1e-9:
        km_priority = "ALTA"
        km_text = f"A distância média aumenta de {_fmt_number(km['atual'])} km para {_fmt_number(km['proposto'])} km."
    elif km_delta < -1e-9:
        km_priority = "POSITIVA"
        km_text = f"A distância média reduz de {_fmt_number(km['atual'])} km para {_fmt_number(km['proposto'])} km."
    else:
        km_priority = "NEUTRA"
        km_text = f"A distância média permanece estável em {_fmt_number(km['proposto'])} km."
    add(
        km_priority,
        "Distância nacional",
        km_text,
        f"Variação de {_fmt_number(km['delta'])} km ({_fmt_number(float(km['delta_pct']) * 100)}%).",
    )
    add(
        "MÉDIA",
        "Cobertura",
        f"O cenário proposto reúne {_fmt_number(stores['proposto'], 0)} correspondentes.",
        f"Diferença de {_fmt_number(stores['delta'], 0)} frente ao atual.",
    )

    if comparison.empty:
        add("ALTA", "Comparabilidade", "Nenhum ID de carteira coincide entre os cenários; a comparação carteira a carteira não é possível.", "As análises nacionais e por gerência de área permanecem válidas de forma independente.")
    else:
        ordered_up = comparison.sort_values(["delta_km_medio", "pole_id"], ascending=[False, True]).head(3)
        ordered_down = comparison.sort_values(["delta_km_medio", "pole_id"], ascending=[True, True]).head(3)
        for _, row in ordered_up.iterrows():
            if pd.notna(row["delta_km_medio"]) and row["delta_km_medio"] > 0:
                add("ALTA", "Carteira", f"{row['polo_proposto']} apresenta uma das maiores altas de distância média.", f"Polo {row['pole_id']}: +{_fmt_number(row['delta_km_medio'])} km.")
        for _, row in ordered_down.iterrows():
            if pd.notna(row["delta_km_medio"]) and row["delta_km_medio"] < 0:
                add("POSITIVA", "Carteira", f"{row['polo_proposto']} apresenta uma das maiores reduções de distância média.", f"Polo {row['pole_id']}: {_fmt_number(row['delta_km_medio'])} km.")
        lowest = comparison.sort_values(["retencao_cobertura_municipal_pct", "pole_id"]).head(3)
        for _, row in lowest.iterrows():
            if pd.notna(row["retencao_cobertura_municipal_pct"]):
                add("MÉDIA", "Retenção territorial", f"{row['polo_proposto']} retém {_fmt_number(row['retencao_cobertura_municipal_pct'] * 100)}% da cobertura municipal atual.", f"{int(row['unidades_removidas'])} unidades exatas removidas e {int(row['unidades_adicionadas'])} adicionadas.")
        shift_metrics = (
            ("correspondentes", "Correspondentes", 0),
            ("populacao", "População", 0),
            ("unidades_territoriais", "Unidades territoriais", 0),
        )
        for metric, label, decimals in shift_metrics:
            delta_column = f"delta_{metric}"
            valid = comparison.dropna(subset=[delta_column]).copy()
            if valid.empty:
                continue
            valid["_variacao_absoluta"] = valid[delta_column].abs()
            largest = valid.sort_values(
                ["_variacao_absoluta", "pole_id"], ascending=[False, True]
            ).iloc[0]
            if largest["_variacao_absoluta"] <= 1e-9:
                continue
            add(
                "MÉDIA",
                label,
                f"{largest['polo_proposto']} concentra a maior variação absoluta de {label.lower()} entre as carteiras correspondentes.",
                f"Polo {largest['pole_id']}: {_fmt_number(largest[f'atual_{metric}'], decimals)} → {_fmt_number(largest[f'proposto_{metric}'], decimals)} (delta {_fmt_number(largest[delta_column], decimals)}).",
            )

    if not area.empty:
        valid_area = area.dropna(subset=["delta_km_medio"])
        if not valid_area.empty:
            worst = valid_area.sort_values(["delta_km_medio", "gerencia_area"], ascending=[False, True]).iloc[0]
            best = valid_area.sort_values(["delta_km_medio", "gerencia_area"], ascending=[True, True]).iloc[0]
            if worst["delta_km_medio"] > 1e-9:
                add("ALTA", "Gerência de área", f"{worst['gerencia_area']} tem o maior impacto de alta na distância média.", f"Variação de {_fmt_number(worst['delta_km_medio'])} km.")
            if best["delta_km_medio"] < -1e-9:
                add("POSITIVA", "Gerência de área", f"{best['gerencia_area']} tem a maior redução de distância média.", f"Variação de {_fmt_number(best['delta_km_medio'])} km.")
            if worst["delta_km_medio"] <= 1e-9 and best["delta_km_medio"] >= -1e-9:
                add("NEUTRA", "Gerência de área", "As distâncias médias permanecem estáveis em todas as gerências de área.", "Nenhuma variação relevante foi identificada no agrupamento superior.")

    if not unmatched.empty:
        old_only = int((unmatched["situacao"] == "SOMENTE_ATUAL").sum())
        new_only = int((unmatched["situacao"] == "SOMENTE_PROPOSTO").sum())
        add("MÉDIA", "IDs não correspondentes", f"Há {old_only} carteiras apenas no cenário atual e {new_only} apenas no proposto.", "Essas carteiras não entram nos deltas carteira a carteira.")
    if not quality.empty:
        errors = int((quality["severidade"] == "ERRO").sum())
        warnings = int((quality["severidade"] == "AVISO").sum())
        add("ALTA" if errors else "MÉDIA", "Qualidade dos dados", f"Foram identificados {errors} erros e {warnings} avisos de qualidade.", "Consulte a aba Qualidade_Dados antes de usar o relatório para decisão final.")
    return pd.DataFrame(records, columns=["prioridade", "tema", "insight", "evidencia"])


def analyze_scenarios(
    current_payload: Mapping[str, Any],
    proposed_payload: Mapping[str, Any],
    current_source: str = "",
    proposed_source: str = "",
) -> AnalysisReport:
    current = prepare_scenario(current_payload, "ATUAL", current_source)
    proposed = prepare_scenario(proposed_payload, "PROPOSTO", proposed_source)
    current_portfolios = _portfolio_frame(current, "ATUAL")
    proposed_portfolios = _portfolio_frame(proposed, "PROPOSTO")
    summary = _summary(current, proposed)
    comparison = _portfolio_comparison(current, proposed, current_portfolios, proposed_portfolios)
    area = _area_comparison(current, proposed)
    movements = _movements(current, proposed)
    unmatched = _unmatched(current, proposed, current_portfolios, proposed_portfolios)
    quality = pd.DataFrame(
        [*current.issues, *proposed.issues],
        columns=["cenario", "severidade", "codigo", "tipo_entidade", "entidade_id", "mensagem"],
    )
    insights = _build_insights(summary, comparison, area, unmatched, quality)
    return AnalysisReport(current, proposed, summary, current_portfolios, proposed_portfolios, comparison, area, movements, unmatched, quality, insights)


def _markdown_table(frame: pd.DataFrame, columns: Sequence[str], limit: int | None = None) -> str:
    selected = frame.loc[:, [column for column in columns if column in frame.columns]]
    if limit is not None:
        selected = selected.head(limit)
    if selected.empty:
        return "_Sem registros._"
    headers = [str(column).replace("_", " ").title() for column in selected.columns]
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for values in selected.itertuples(index=False, name=None):
        cells = []
        for value in values:
            if isinstance(value, float):
                cells.append(_fmt_number(value))
            else:
                cells.append(str(value).replace("|", "\\|"))
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def build_markdown(report: AnalysisReport) -> str:
    lines = [
        "# Análise comparativa de cenários territoriais",
        "",
        f"- Cenário atual: **{report.current.name}**",
        f"- Cenário proposto: **{report.proposed.name}**",
        f"- Gerado em: **{datetime.now().astimezone().strftime('%d/%m/%Y %H:%M:%S %Z')}**",
        "",
        "## Resumo nacional",
        "",
        _markdown_table(report.summary, ["indicador", "atual", "proposto", "delta", "delta_pct"]),
        "",
        "## Principais insights",
        "",
    ]
    for row in report.insights.itertuples(index=False):
        lines.append(f"- **[{row.prioridade}] {row.tema}:** {row.insight} {row.evidencia}")
    lines.extend(
        [
            "",
            "## Impacto por gerência de área",
            "",
            _markdown_table(report.area_comparison, ["gerencia_area", "atual_carteiras", "proposto_carteiras", "delta_carteiras", "atual_km_medio", "proposto_km_medio", "delta_km_medio"]),
            "",
            "## Carteiras com maior aumento de distância média",
            "",
            _markdown_table(report.portfolio_comparison.sort_values("delta_km_medio", ascending=False) if not report.portfolio_comparison.empty else report.portfolio_comparison, ["pole_id", "polo_proposto", "atual_km_medio", "proposto_km_medio", "delta_km_medio", "retencao_cobertura_municipal_pct"], 10),
            "",
            "## Limitações e qualidade",
            "",
            "- A correspondência de carteiras usa exclusivamente o mesmo `pole_id` nos dois cenários.",
            "- As distâncias são recalculadas por Haversine; o campo `distanceKm` do JSON não é usado.",
            "- Municípios repetidos dentro da mesma carteira têm correspondentes somados e população máxima; distritos permanecem unidades separadas.",
            "- Agregações por gerência de área são calculadas diretamente sobre as unidades, sem média de médias.",
            f"- Ocorrências de qualidade: **{len(report.quality)}**.",
            "",
        ]
    )
    if report.portfolio_comparison.empty:
        lines.append("> Não há IDs de carteira em comum. Por isso, os cenários são apresentados de forma independente nos níveis nacional e de gerência de área.")
        lines.append("")
    return "\n".join(lines)


def _style_workbook(path: Path, report: AnalysisReport) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:
        raise RuntimeError("A exportação Excel requer openpyxl. Instale requirements-v5.txt ou requirements-v4.txt.") from exc

    wb = load_workbook(path)
    navy, teal, pale, white = "16324F", "2D7D7A", "E8F1F5", "FFFFFF"
    thin = Side(style="thin", color="D9E2E8")
    for position, ws in enumerate(wb.worksheets, start=1):
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A6" if ws.title == "Resumo" else "A2"
        ws.auto_filter.ref = ws.dimensions
        header_row = 5 if ws.title == "Resumo" else 1
        if ws.max_column:
            for cell in ws[header_row]:
                cell.fill = PatternFill("solid", fgColor=navy)
                cell.font = Font(color=white, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[header_row].height = 32
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(max_length + 2, 11), 48)
        for row in ws.iter_rows(min_row=header_row + 1):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=ws.title in {"Insights", "Qualidade_Dados"})
                if isinstance(cell.value, float):
                    cell.number_format = "0.00"
                cell.border = Border(bottom=thin)
        if ws.max_row > header_row and ws.max_column > 0:
            reference = f"A{header_row}:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            table_name = re.sub(r"[^A-Za-z0-9_]", "_", f"Tabela_{position}_{ws.title}")[:200]
            table = Table(displayName=table_name, ref=reference)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
        headers = {str(ws.cell(header_row, col).value): col for col in range(1, ws.max_column + 1)}
        for name, col in headers.items():
            if name.endswith("_pct") or name == "delta_pct":
                for row in range(header_row + 1, ws.max_row + 1):
                    ws.cell(row, col).number_format = "0.0%"
            elif "km" in name.lower():
                for row in range(header_row + 1, ws.max_row + 1):
                    ws.cell(row, col).number_format = '0.00 "km"'
            elif name.startswith("delta_") and ws.max_row > header_row:
                cell_range = f"{ws.cell(header_row + 1, col).coordinate}:{ws.cell(ws.max_row, col).coordinate}"
                ws.conditional_formatting.add(cell_range, ColorScaleRule(start_type="min", start_color="F4CCCC", mid_type="num", mid_value=0, mid_color="FFFFFF", end_type="max", end_color="D9EAD3"))
    summary_ws = wb["Resumo"]
    summary_ws.merge_cells("A1:D1")
    summary_ws["A1"] = "Análise comparativa de cenários territoriais"
    summary_ws["A1"].fill = PatternFill("solid", fgColor=navy)
    summary_ws["A1"].font = Font(color=white, bold=True, size=16)
    summary_ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    summary_ws.row_dimensions[1].height = 28
    summary_ws["A2"] = "Cenário atual"
    summary_ws["B2"] = report.current.name
    summary_ws["A3"] = "Cenário proposto"
    summary_ws["B3"] = report.proposed.name
    summary_ws["A2"].font = summary_ws["A3"].font = Font(bold=True, color=teal)
    summary_ws["A2"].fill = summary_ws["A3"].fill = PatternFill("solid", fgColor=pale)
    indicator_rows = {summary_ws.cell(row, 1).value: row for row in range(6, summary_ws.max_row + 1)}
    chart_rows = [
        indicator_rows[name]
        for name in ("Km médio", "Km mediano", "Km P90", "Km máximo")
        if name in indicator_rows
    ]
    if chart_rows and max(chart_rows) - min(chart_rows) == len(chart_rows) - 1:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Distâncias: atual x proposto"
        chart.y_axis.title = "Quilômetros"
        data = Reference(summary_ws, min_col=2, max_col=3, min_row=min(chart_rows) - 1, max_row=max(chart_rows))
        categories = Reference(summary_ws, min_col=1, min_row=min(chart_rows), max_row=max(chart_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.height = 7
        chart.width = 12
        summary_ws.add_chart(chart, "G2")
    wb.save(path)


def export_report(report: AnalysisReport, output_dir: str | Path, timestamp: datetime | None = None) -> tuple[Path, Path]:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    stamp = (timestamp or datetime.now(timezone.utc).astimezone()).strftime("%Y%m%d_%H%M%S")
    excel_path = output / f"analise_cenarios_{stamp}.xlsx"
    markdown_path = output / f"analise_cenarios_{stamp}.md"
    sheets = {
        "Resumo": report.summary,
        "Carteiras_Atual": report.current_portfolios,
        "Carteiras_Propostas": report.proposed_portfolios,
        "Comparacao_Carteiras": report.portfolio_comparison,
        "Gerencias_Area": report.area_comparison,
        "Movimentacoes": report.movements,
        "Nao_Correspondentes": report.unmatched,
        "Qualidade_Dados": report.quality,
        "Insights": report.insights,
    }
    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            for sheet_name, frame in sheets.items():
                startrow = 4 if sheet_name == "Resumo" else 0
                frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)
    except ImportError as exc:
        raise RuntimeError("A exportação Excel requer openpyxl. Instale requirements-v5.txt ou requirements-v4.txt.") from exc
    _style_workbook(excel_path, report)
    markdown_path.write_text(build_markdown(report), encoding="utf-8")
    return excel_path, markdown_path


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compara um JSON exportado pelo Builder com o cenário territorial atual.")
    parser.add_argument("--builder-json", required=True, help="JSON exportado pelo Builder (envelope ou ScenarioData bruto).")
    parser.add_argument("--current-json", default=".territorios-data/current.json", help="JSON do cenário atual. Padrão: .territorios-data/current.json")
    parser.add_argument("--output-dir", default="analise_builder", help="Diretório dos relatórios. Padrão: analise_builder")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    current_path, proposed_path = Path(args.current_json), Path(args.builder_json)
    if not current_path.is_file():
        raise SystemExit(f"Cenário atual não encontrado: {current_path}")
    if not proposed_path.is_file():
        raise SystemExit(f"JSON do Builder não encontrado: {proposed_path}")
    try:
        report = analyze_scenarios(
            load_scenario(current_path),
            load_scenario(proposed_path),
            str(current_path.resolve()),
            str(proposed_path.resolve()),
        )
        excel_path, markdown_path = export_report(report, args.output_dir)
    except (TypeError, ValueError, RuntimeError) as exc:
        raise SystemExit(f"Falha na análise: {exc}") from exc
    print(f"Excel: {excel_path.resolve()}")
    print(f"Markdown: {markdown_path.resolve()}")
    print(f"Ocorrências de qualidade: {len(report.quality)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())