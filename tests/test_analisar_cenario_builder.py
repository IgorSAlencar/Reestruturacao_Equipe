import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import pytest

import Analisar_Cenario_Builder as analyzer


def pole(pole_id: str, lon: float, lat: float = 0.0, area: str = "AREA A") -> dict:
    return {
        "id": pole_id,
        "name": f"Polo {pole_id}",
        "lon": lon,
        "lat": lat,
        "area": area,
        "regional": "Regional 1",
        "uf": "SP",
    }


def unit(
    unit_id: str,
    municipality_code: str,
    pole_id: str,
    lon: float,
    *,
    stores: float = 1,
    population: float = 100,
    unit_type: str = "MUNICIPALITY",
    district_code: str | None = None,
) -> dict:
    return {
        "id": unit_id,
        "type": unit_type,
        "municipalityCode": municipality_code,
        "districtCode": district_code,
        "name": unit_id,
        "uf": "SP",
        "poleId": pole_id,
        "population": population,
        "stores": stores,
        "lat": 0.0,
        "lon": lon,
        "distanceKm": 0,
    }


def scenario(poles: list[dict], units: list[dict], title: str = "Cenário") -> dict:
    return {
        "id": title.lower().replace(" ", "-"),
        "title": title,
        "summary": {
            "totalPoles": len(poles),
            "totalUnits": len(units),
            "totalStores": sum(item["stores"] for item in units),
        },
        "poles": poles,
        "units": units,
    }


def test_accepts_raw_scenario_and_builder_envelope(tmp_path: Path) -> None:
    raw = scenario([pole("1", 0)], [unit("U1", "100", "1", 1)])
    envelope = {"id": "draft-1", "revision": 3, "data": raw}
    raw_path = tmp_path / "raw.json"
    envelope_path = tmp_path / "envelope.json"
    raw_path.write_text(json.dumps(raw), encoding="utf-8")
    envelope_path.write_text(json.dumps(envelope), encoding="utf-8")

    assert analyzer.load_scenario(raw_path)["poles"][0]["id"] == "1"
    assert analyzer.load_scenario(envelope_path)["units"][0]["id"] == "U1"


def test_haversine_recalculates_distance_instead_of_json_value() -> None:
    assert analyzer.haversine_km(0, 0, 0, 1) == pytest.approx(111.195, rel=1e-4)
    payload = scenario([pole("1", 0)], [unit("U1", "100", "1", 1)])
    prepared = analyzer.prepare_scenario(payload, "ATUAL")
    assert prepared.rows[0]["distancia_km"] == pytest.approx(111.195, rel=1e-4)


def test_accepts_long_coordinate_and_municipality_name_fields_used_by_current_json() -> None:
    payload = scenario([pole("1", 0)], [unit("U1", "100", "1", 1)])
    payload["poles"][0]["latitude"] = payload["poles"][0].pop("lat")
    payload["poles"][0]["longitude"] = payload["poles"][0].pop("lon")
    payload["units"][0]["latitude"] = payload["units"][0].pop("lat")
    payload["units"][0]["longitude"] = payload["units"][0].pop("lon")
    payload["units"][0]["municipalityName"] = "Cidade real"
    payload["units"][0].pop("name")
    prepared = analyzer.prepare_scenario(payload, "ATUAL")
    assert prepared.rows[0]["distancia_km"] == pytest.approx(111.195, rel=1e-4)
    assert prepared.rows[0]["nome"] == "Cidade real"


def test_deduplicates_municipality_and_calculates_weighted_means() -> None:
    payload = scenario(
        [pole("1", 0)],
        [
            unit("U1", "100", "1", 1, stores=2, population=100),
            unit("U2", "100", "1", 1, stores=3, population=90),
            unit("U3", "200", "1", 2, stores=1, population=200),
        ],
    )
    prepared = analyzer.prepare_scenario(payload, "ATUAL")
    assert len(prepared.rows) == 2
    first = next(row for row in prepared.rows if row["municipio_codigo"] == "100")
    assert first["correspondentes"] == 5
    assert first["populacao"] == 100
    metrics = analyzer.calculate_metrics(prepared.rows)
    distance_1 = analyzer.haversine_km(0, 0, 0, 1)
    distance_2 = analyzer.haversine_km(0, 0, 0, 2)
    assert metrics["km_medio"] == pytest.approx((distance_1 + distance_2) / 2)
    assert metrics["km_medio_ponderado_correspondentes"] == pytest.approx((distance_1 * 5 + distance_2) / 6)
    assert metrics["km_medio_ponderado_populacao"] == pytest.approx((distance_1 * 100 + distance_2 * 200) / 300)


def test_matching_is_only_by_same_pole_id_and_lists_unmatched() -> None:
    current = scenario(
        [pole("1", 0), pole("2", 2)],
        [unit("U1", "100", "1", 1), unit("U2", "200", "2", 3)],
        "Atual",
    )
    proposed = scenario(
        [pole("1", 0), pole("3", 2)],
        [unit("U1", "100", "1", 0.5), unit("U2", "200", "3", 2.5)],
        "Proposto",
    )
    report = analyzer.analyze_scenarios(current, proposed)
    assert report.portfolio_comparison["pole_id"].tolist() == ["1"]
    assert set(report.unmatched["situacao"]) == {"SOMENTE_ATUAL", "SOMENTE_PROPOSTO"}
    assert set(report.unmatched["pole_id"]) == {"2", "3"}


def test_district_conversion_preserves_municipal_coverage_but_not_exact_unit() -> None:
    current = scenario([pole("1", 0)], [unit("M1", "100", "1", 1)], "Atual")
    proposed = scenario(
        [pole("1", 0)],
        [unit("D1", "100", "1", 1, unit_type="DISTRICT", district_code="10001")],
        "Proposto",
    )
    comparison = analyzer.analyze_scenarios(current, proposed).portfolio_comparison.iloc[0]
    assert comparison["retencao_exata_pct"] == 0
    assert comparison["retencao_cobertura_municipal_pct"] == 1
    assert comparison["unidades_removidas"] == 1
    assert comparison["unidades_adicionadas"] == 1


def test_area_rollup_uses_units_directly_and_movement_uses_dominant_origin() -> None:
    current = scenario(
        [pole("1", 0, area="AREA A"), pole("2", 4, area="AREA A")],
        [
            unit("U1", "100", "1", 1, stores=10),
            unit("U2", "100", "2", 1, stores=5),
            unit("U3", "200", "2", 2, stores=1),
        ],
        "Atual",
    )
    proposed = scenario(
        [pole("1", 0, area="AREA A"), pole("2", 4, area="AREA A")],
        [unit("U1", "100", "2", 1, stores=15), unit("U3", "200", "2", 2, stores=1)],
        "Proposto",
    )
    report = analyzer.analyze_scenarios(current, proposed)
    movement = report.movements.loc[report.movements["cobertura_municipal_chave"] == "MUN:100"].iloc[0]
    assert movement["polo_origem_id"] == "1"
    assert movement["polo_destino_id"] == "2"
    assert movement["tipo_movimento"] == "MOVIDA"

    current_prepared = analyzer.prepare_scenario(current, "ATUAL")
    expected = sum(row["distancia_km"] for row in current_prepared.rows) / len(current_prepared.rows)
    area = report.area_comparison.loc[report.area_comparison["gerencia_area"] == "AREA A"].iloc[0]
    assert area["atual_km_medio"] == pytest.approx(expected)


def test_dominant_origin_tie_uses_lowest_pole_id() -> None:
    current = scenario(
        [pole("10", 0), pole("2", 2)],
        [unit("U1", "100", "10", 1, stores=5), unit("U2", "100", "2", 1, stores=5)],
    )
    proposed = scenario([pole("10", 0)], [unit("U1", "100", "10", 1, stores=5)])
    movement = analyzer.analyze_scenarios(current, proposed).movements.iloc[0]
    assert movement["polo_origem_id"] == "2"


def test_no_matching_ids_keeps_independent_area_analysis_and_insight() -> None:
    current = scenario([pole("1", 0, area="AREA A")], [unit("U1", "100", "1", 1)])
    proposed = scenario([pole("2", 0, area="AREA A")], [unit("U1", "100", "2", 0.5)])
    report = analyzer.analyze_scenarios(current, proposed)
    assert report.portfolio_comparison.empty
    assert not report.area_comparison.empty
    assert report.insights["insight"].str.contains("Nenhum ID de carteira").any()


def test_zero_distance_delta_is_reported_as_stable() -> None:
    payload = scenario([pole("1", 0)], [unit("U1", "100", "1", 1)])
    report = analyzer.analyze_scenarios(payload, payload)
    national = report.insights.loc[report.insights["tema"] == "Distância nacional"].iloc[0]
    assert national["prioridade"] == "NEUTRA"
    assert "permanece estável" in national["insight"]
    assert report.insights["insight"].str.contains("estáveis em todas as gerências").any()


def test_insights_include_largest_store_population_and_unit_shifts() -> None:
    current = scenario(
        [pole("1", 0)],
        [unit("U1", "100", "1", 1, stores=1, population=100)],
    )
    proposed = scenario(
        [pole("1", 0)],
        [
            unit("U1", "100", "1", 1, stores=3, population=200),
            unit("U2", "200", "1", 2, stores=2, population=50),
        ],
    )
    report = analyzer.analyze_scenarios(current, proposed)
    assert {"Correspondentes", "População", "Unidades territoriais"} <= set(
        report.insights["tema"]
    )


def test_quality_flags_missing_poles_unassigned_units_and_invalid_coordinates() -> None:
    payload = scenario(
        [pole("1", 0)],
        [
            unit("U1", "100", "999", 1),
            unit("U2", "200", "", 1),
            unit("U3", "300", "1", 999),
        ],
    )
    prepared = analyzer.prepare_scenario(payload, "ATUAL")
    codes = {issue["codigo"] for issue in prepared.issues}
    assert {"POLO_NAO_ENCONTRADO", "UNIDADE_SEM_POLO", "COORDENADA_UNIDADE_INVALIDA"} <= codes


def test_exports_expected_excel_sheets_and_markdown(tmp_path: Path) -> None:
    current = scenario([pole("1", 0)], [unit("U1", "100", "1", 1)], "Atual")
    proposed = scenario([pole("1", 0)], [unit("U1", "100", "1", 0.5)], "Proposto")
    report = analyzer.analyze_scenarios(current, proposed)
    excel_path, markdown_path = analyzer.export_report(
        report, tmp_path, datetime(2026, 8, 13, 12, 30, 0, tzinfo=timezone.utc)
    )

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, read_only=True, data_only=False)
    assert workbook.sheetnames == [
        "Resumo",
        "Carteiras_Atual",
        "Carteiras_Propostas",
        "Comparacao_Carteiras",
        "Gerencias_Area",
        "Movimentacoes",
        "Nao_Correspondentes",
        "Qualidade_Dados",
        "Insights",
    ]
    assert "Análise comparativa" in workbook["Resumo"]["A1"].value
    markdown = markdown_path.read_text(encoding="utf-8")
    assert "## Resumo nacional" in markdown
    assert "## Principais insights" in markdown
    assert "Haversine" in markdown


def test_cli_generates_both_reports(tmp_path: Path) -> None:
    current = scenario([pole("1", 0)], [unit("U1", "100", "1", 1)], "Atual")
    proposed = {"id": "draft", "revision": 1, "data": current}
    current_path = tmp_path / "current.json"
    builder_path = tmp_path / "builder.json"
    output_path = tmp_path / "output"
    current_path.write_text(json.dumps(current), encoding="utf-8")
    builder_path.write_text(json.dumps(proposed), encoding="utf-8")

    result = analyzer.main(
        [
            "--builder-json",
            str(builder_path),
            "--current-json",
            str(current_path),
            "--output-dir",
            str(output_path),
        ]
    )
    assert result == 0
    assert len(list(output_path.glob("*.xlsx"))) == 1
    assert len(list(output_path.glob("*.md"))) == 1


def test_real_current_json_and_latest_sqlite_draft_when_available() -> None:
    root = Path(__file__).resolve().parents[1]
    current_path = root / ".territorios-data" / "current.json"
    database_path = root / ".territorios-data" / "territorios.sqlite"
    if not current_path.is_file() or not database_path.is_file():
        pytest.skip("Dados locais da aplicação não estão disponíveis.")
    with sqlite3.connect(database_path) as connection:
        row = connection.execute("SELECT data_json FROM drafts ORDER BY updated_at DESC LIMIT 1").fetchone()
    if row is None:
        pytest.skip("O banco local não contém rascunho do Builder.")
    report = analyzer.analyze_scenarios(
        analyzer.load_scenario(current_path),
        json.loads(row[0]),
        str(current_path),
        str(database_path),
    )
    assert len(report.current.poles) > 0
    assert len(report.proposed.poles) > 0
    assert not report.summary.empty
    assert not report.area_comparison.empty
